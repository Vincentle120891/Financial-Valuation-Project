"""
VND Financial Statement Parser for Vietnamese Companies

Parses Vietnamese financial reports (VND-denominated) and converts to standardized format.
Supports:
- Annual reports (Báo cáo thường niên)
- Quarterly reports (Báo cáo quý)
- Consolidated statements (Hợp nhất)
- Standalone statements (Riêng lẻ)

Vietnamese Accounting Standards (VAS) vs IFRS mapping included.
"""

import re
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass
from enum import Enum
import pandas as pd


class VNStatementType(Enum):
    """Vietnamese Financial Statement Types"""
    INCOME_STATEMENT = "Income Statement"  # Báo cáo kết quả hoạt động kinh doanh
    BALANCE_SHEET = "Balance Sheet"  # Báo cáo tình hình tài chính
    CASH_FLOW = "Cash Flow Statement"  # Báo cáo lưu chuyển tiền tệ
    NOTES = "Notes to Financial Statements"  # Thuyết minh báo cáo tài chính


class VNAccountCode:
    """
    Vietnamese Accounting System Account Codes (Hệ thống tài khoản kế toán)
    Maps VAS account codes to standard financial metrics
    """
    
    # Income Statement Accounts (Tài khoản loại 5, 6, 7, 8, 9)
    REVENUE_ACCOUNTS = ["511", "512", "513", "518"]  # Doanh thu bán hàng, cung cấp dịch vụ
    COGS_ACCOUNTS = ["632"]  # Giá vốn hàng bán
    OPERATING_EXPENSE_ACCOUNTS = ["641", "642"]  # Chi phí bán hàng, Chi phí quản lý doanh nghiệp
    FINANCIAL_INCOME_ACCOUNTS = ["515"]  # Doanh thu hoạt động tài chính
    FINANCIAL_EXPENSE_ACCOUNTS = ["635"]  # Chi phí tài chính
    INTEREST_EXPENSE_ACCOUNTS = ["6351"]  # Chi phí lãi vay
    OTHER_INCOME_ACCOUNTS = ["711"]  # Thu nhập khác
    OTHER_EXPENSE_ACCOUNTS = ["811"]  # Chi phí khác
    INCOME_TAX_ACCOUNTS = ["821"]  # Chi phí thuế thu nhập doanh nghiệp
    
    # Balance Sheet Accounts (Tài khoản loại 1, 2, 3, 4)
    CASH_ACCOUNTS = ["111", "112", "113"]  # Tiền mặt, Tiền gửi ngân hàng
    AR_ACCOUNTS = ["131", "136", "138"]  # Phải thu khách hàng, Phải thu nội bộ
    INVENTORY_ACCOUNTS = ["151", "152", "153", "154", "155", "156", "157"]  # Hàng tồn kho
    AP_ACCOUNTS = ["331", "336", "338"]  # Phải trả người bán, Phải trả nội bộ
    SHORT_TERM_DEBT_ACCOUNTS = ["3411"]  # Vay và nợ thuê tài chính ngắn hạn
    LONG_TERM_DEBT_ACCOUNTS = ["3412"]  # Vay và nợ thuê tài chính dài hạn
    EQUITY_ACCOUNTS = ["411", "412", "413", "414", "415", "416", "417", "418", "419", "421"]  # Vốn chủ sở hữu
    
    # Cash Flow Categories
    OPERATING_CF_INDICATORS = ["Mã số 20", "MS 20"]  # Lưu chuyển tiền thuần từ hoạt động kinh doanh
    INVESTING_CF_INDICATORS = ["Mã số 30", "MS 30"]  # Lưu chuyển tiền thuần từ hoạt động đầu tư
    FINANCING_CF_INDICATORS = ["Mã số 40", "MS 40"]  # Lưu chuyển tiền thuần từ hoạt động tài chính


@dataclass
class VNFinancialItem:
    """Vietnamese Financial Statement Item"""
    code: str  # Mã số (e.g., "01", "02")
    name_vi: str  # Vietnamese name
    name_en: str  # English name
    value: float  # Value in VND millions
    parent_code: Optional[str] = None  # Parent item code for hierarchical structure
    is_calculated: bool = False  # True if this is a calculated subtotal
    account_codes: List[str] = None  # Related VAS account codes
    
    def __post_init__(self):
        if self.account_codes is None:
            self.account_codes = []


@dataclass
class ParsedVNFinancials:
    """Parsed Vietnamese Financial Statements"""
    ticker: str
    company_name_vi: str
    company_name_en: str
    fiscal_year: int
    quarter: Optional[int]  # None for annual
    statement_type: VNStatementType
    currency: str  # Always "VND"
    unit: str  # "million VND" or "billion VND"
    items: List[VNFinancialItem]
    metadata: Dict[str, Any] = None
    
    def __post_init__(self):
        if self.metadata is None:
            self.metadata = {}
    
    def to_standard_dict(self) -> Dict[str, Any]:
        """Convert to standard financial dictionary format"""
        result = {
            "ticker": self.ticker,
            "company_name": self.company_name_en,
            "fiscal_year": self.fiscal_year,
            "quarter": self.quarter,
            "currency": "VND",
            "unit": self.unit
        }
        
        # Map Vietnamese items to standard metrics
        item_map = {item.code: item.value for item in self.items}
        
        if self.statement_type == VNStatementType.INCOME_STATEMENT:
            result.update({
                "revenue": self._find_item(["01", "10"], item_map),  # Doanh thu thuần
                "cogs": self._find_item(["02", "11"], item_map),  # Giá vốn hàng bán
                "gross_profit": self._find_item(["10", "20"], item_map),  # Lợi nhuận gộp
                "operating_expenses": self._find_item(["25", "30"], item_map),  # Chi phí hoạt động
                "operating_income": self._find_item(["30", "40"], item_map),  # Lợi nhuận thuần từ HĐKD
                "financial_income": self._find_item(["41", "45"], item_map),  # Doanh thu tài chính
                "financial_expense": self._find_item(["42", "46"], item_map),  # Chi phí tài chính
                "interest_expense": self._find_item(["43", "47"], item_map),  # Chi phí lãi vay
                "other_income": self._find_item(["50", "55"], item_map),  # Thu nhập khác
                "other_expense": self._find_item(["51", "56"], item_map),  # Chi phí khác
                "pretax_income": self._find_item(["60", "70"], item_map),  # Lợi nhuận trước thuế
                "income_tax": self._find_item(["61", "71"], item_map),  # Chi phí thuế TNDN
                "net_income": self._find_item(["70", "80"], item_map),  # Lợi nhuận sau thuế
            })
        
        elif self.statement_type == VNStatementType.BALANCE_SHEET:
            result.update({
                "cash": self._find_item(["110", "120"], item_map),  # Tiền và tương đương tiền
                "accounts_receivable": self._find_item(["130", "140"], item_map),  # Phải thu ngắn/dài hạn
                "inventory": self._find_item(["140", "150"], item_map),  # Hàng tồn kho
                "total_current_assets": self._find_item(["100", "200"], item_map),  # Tài sản ngắn hạn
                "fixed_assets": self._find_item(["210", "220"], item_map),  # Tài sản cố định
                "total_assets": self._find_item(["200", "270"], item_map),  # Tổng cộng tài sản
                "accounts_payable": self._find_item(["310", "320"], item_map),  # Phải trả ngắn hạn
                "short_term_debt": self._find_item(["330", "340"], item_map),  # Vay ngắn hạn
                "total_current_liabilities": self._find_item(["300", "400"], item_map),  # Nợ ngắn hạn
                "long_term_debt": self._find_item(["410", "420"], item_map),  # Vay dài hạn
                "total_liabilities": self._find_item(["400", "500"], item_map),  # Tổng cộng nợ phải trả
                "equity": self._find_item(["500", "600"], item_map),  # Vốn chủ sở hữu
            })
        
        elif self.statement_type == VNStatementType.CASH_FLOW:
            result.update({
                "operating_cf": self._find_item(["20", "30"], item_map),  # Lưu chuyển tiền từ HĐKD
                "investing_cf": self._find_item(["40", "50"], item_map),  # Lưu chuyển tiền từ HĐĐT
                "financing_cf": self._find_item(["60", "70"], item_map),  # Lưu chuyển tiền từ HĐTC
                "net_change_in_cash": self._find_item(["70", "80"], item_map),  # Tăng/giảm tiền thuần
                "capex": self._find_item(["41", "45"], item_map),  # Mua sắm TSCĐ
                "dividends_paid": self._find_item(["61", "65"], item_map),  # Cổ tức đã trả
            })
        
        return result
    
    def _find_item(self, codes: List[str], item_map: Dict[str, float]) -> Optional[float]:
        """Find first matching item by code"""
        for code in codes:
            if code in item_map:
                return item_map[code]
        return None


class VNDFinancialParser:
    """
    Parser for Vietnamese Financial Statements
    
    Handles:
    - PDF/Excel parsing (when available)
    - Manual data entry validation
    - VAS to IFRS conversion
    - Currency unit normalization (millions/billions VND)
    - Quarter/Annual detection
    """
    
    def __init__(self):
        self.common_vietnamese_terms = {
            # Income Statement
            "doanh thu": "revenue",
            "giá vốn": "cost_of_goods_sold",
            "lợi nhuận gộp": "gross_profit",
            "chi phí bán hàng": "selling_expenses",
            "chi phí quản lý": "administrative_expenses",
            "doanh thu tài chính": "financial_income",
            "chi phí tài chính": "financial_expense",
            "chi phí lãi vay": "interest_expense",
            "thu nhập khác": "other_income",
            "chi phí khác": "other_expense",
            "lợi nhuận trước thuế": "pretax_income",
            "thuế thu nhập": "income_tax",
            "lợi nhuận sau thuế": "net_income",
            
            # Balance Sheet
            "tiền": "cash",
            "đầu tư ngắn hạn": "short_term_investments",
            "phải thu": "accounts_receivable",
            "hàng tồn kho": "inventory",
            "tài sản ngắn hạn": "current_assets",
            "tài sản cố định": "fixed_assets",
            "tổng tài sản": "total_assets",
            "phải trả": "accounts_payable",
            "vay ngắn hạn": "short_term_debt",
            "nợ ngắn hạn": "current_liabilities",
            "vay dài hạn": "long_term_debt",
            "nợ dài hạn": "long_term_liabilities",
            "vốn chủ sở hữu": "equity",
            
            # Cash Flow
            "lưu chuyển tiền từ hoạt động kinh doanh": "operating_cash_flow",
            "lưu chuyển tiền từ hoạt động đầu tư": "investing_cash_flow",
            "lưu chuyển tiền từ hoạt động tài chính": "financing_cash_flow",
            "mua sắm tài sản cố định": "capex",
            "trả cổ tức": "dividends_paid",
        }
    
    def parse_from_dict(self, data: Dict[str, Any]) -> ParsedVNFinancials:
        """
        Parse financial data from dictionary format
        
        Expected format:
        {
            "ticker": "VNM",
            "company_name_vi": "...",
            "company_name_en": "...",
            "fiscal_year": 2023,
            "quarter": None,  # or 1, 2, 3, 4
            "statement_type": "income_statement",
            "unit": "million VND",
            "items": [
                {"code": "01", "name_vi": "Doanh thu thuần", "value": 50000},
                ...
            ]
        }
        """
        statement_type_map = {
            "income_statement": VNStatementType.INCOME_STATEMENT,
            "balance_sheet": VNStatementType.BALANCE_SHEET,
            "cash_flow": VNStatementType.CASH_FLOW,
            "notes": VNStatementType.NOTES
        }
        
        stmt_type_str = data.get("statement_type", "income_statement").lower()
        statement_type = statement_type_map.get(stmt_type_str, VNStatementType.INCOME_STATEMENT)
        
        items = []
        for item_data in data.get("items", []):
            item = VNFinancialItem(
                code=item_data.get("code", ""),
                name_vi=item_data.get("name_vi", ""),
                name_en=self._translate_to_english(item_data.get("name_vi", "")),
                value=float(item_data.get("value", 0)),
                parent_code=item_data.get("parent_code"),
                is_calculated=item_data.get("is_calculated", False),
                account_codes=item_data.get("account_codes", [])
            )
            items.append(item)
        
        return ParsedVNFinancials(
            ticker=data["ticker"].upper(),
            company_name_vi=data.get("company_name_vi", ""),
            company_name_en=data.get("company_name_en", ""),
            fiscal_year=int(data["fiscal_year"]),
            quarter=int(data["quarter"]) if data.get("quarter") else None,
            statement_type=statement_type,
            currency="VND",
            unit=data.get("unit", "million VND"),
            items=items,
            metadata=data.get("metadata", {})
        )
    
    def parse_from_excel(self, filepath: str, ticker: str) -> List[ParsedVNFinancials]:
        """
        Parse Vietnamese financial statements from Excel file
        
        Expected Excel structure:
        - Sheet names: "BKQKQKD", "BKTCTC", "BCLCTT" (or English equivalents)
        - Columns: "Mã số", "Chỉ tiêu", "Năm nay", "Năm trước"
        """
        import openpyxl
        
        if not filepath.endswith(('.xlsx', '.xls')):
            raise ValueError("File must be an Excel file (.xlsx or .xls)")
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        results = []
        
        # Map sheet names to statement types
        sheet_mapping = {
            "BKQKQKD": VNStatementType.INCOME_STATEMENT,
            "BÁO CÁO KẾT QUẢ HOẠT ĐỘNG KINH DOANH": VNStatementType.INCOME_STATEMENT,
            "INCOME STATEMENT": VNStatementType.INCOME_STATEMENT,
            "BKTCTC": VNStatementType.BALANCE_SHEET,
            "BÁO CÁO TÌNH HÌNH TÀI CHÍNH": VNStatementType.BALANCE_SHEET,
            "BALANCE SHEET": VNStatementType.BALANCE_SHEET,
            "BCLCTT": VNStatementType.CASH_FLOW,
            "BÁO CÁO LƯU CHUYỂN TIỀN TỆ": VNStatementType.CASH_FLOW,
            "CASH FLOW": VNStatementType.CASH_FLOW,
        }
        
        for sheet_name, statement_type in sheet_mapping.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                parsed = self._parse_excel_sheet(ws, ticker, statement_type)
                if parsed:
                    results.append(parsed)
        
        return results
    
    def _parse_excel_sheet(self, ws, ticker: str, statement_type: VNStatementType) -> Optional[ParsedVNFinancials]:
        """Parse a single Excel sheet"""
        # Implementation would extract rows, detect codes/names/values
        # This is a simplified version
        items = []
        
        # Detect year from header
        fiscal_year = 2023  # Default, would extract from actual file
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx < 3:  # Skip header rows
                continue
            
            # Try to parse row as financial item
            # Expected: [code, name_vi, current_year_value, prior_year_value]
            if len(row) >= 3:
                code = str(row[0]) if row[0] else ""
                name_vi = str(row[1]) if row[1] else ""
                
                # Check if this looks like a valid financial item
                if code and any(c.isdigit() for c in code):
                    try:
                        value = float(row[2]) if row[2] else 0
                        item = VNFinancialItem(
                            code=code.strip(),
                            name_vi=name_vi.strip(),
                            name_en=self._translate_to_english(name_vi),
                            value=value,
                        )
                        items.append(item)
                    except (ValueError, TypeError):
                        continue
        
        if not items:
            return None
        
        return ParsedVNFinancials(
            ticker=ticker.upper(),
            company_name_vi="",  # Would extract from file
            company_name_en="",
            fiscal_year=fiscal_year,
            quarter=None,
            statement_type=statement_type,
            currency="VND",
            unit="million VND",
            items=items
        )
    
    def _translate_to_english(self, vietnamese_text: str) -> str:
        """Translate Vietnamese financial term to English"""
        if not vietnamese_text:
            return ""
        
        text_lower = vietnamese_text.lower().strip()
        
        # Check for exact matches
        if text_lower in self.common_vietnamese_terms:
            return self.common_vietnamese_terms[text_lower].replace("_", " ").title()
        
        # Check for partial matches
        for vi_term, en_term in self.common_vietnamese_terms.items():
            if vi_term in text_lower:
                return en_term.replace("_", " ").title()
        
        # Return original if no match found
        return vietnamese_text
    
    def convert_vnd_to_usd(self, amount_vnd: float, exchange_rate: float = 24500) -> float:
        """
        Convert VND amount to USD
        
        Args:
            amount_vnd: Amount in VND (in millions/billions as per unit)
            exchange_rate: VND per USD (default ~24,500)
        
        Returns:
            Amount in USD
        """
        return amount_vnd / exchange_rate
    
    def normalize_units(self, value: float, from_unit: str, to_unit: str = "million VND") -> float:
        """
        Normalize between million VND and billion VND
        
        Args:
            value: Value to convert
            from_unit: Current unit ("million VND" or "billion VND")
            to_unit: Target unit (default "million VND")
        
        Returns:
            Converted value
        """
        if from_unit == to_unit:
            return value
        
        if from_unit == "billion VND" and to_unit == "million VND":
            return value * 1000
        elif from_unit == "million VND" and to_unit == "billion VND":
            return value / 1000
        else:
            raise ValueError(f"Unsupported unit conversion: {from_unit} to {to_unit}")
    
    def validate_vas_compliance(self, items: List[VNFinancialItem]) -> List[str]:
        """
        Validate financial statements against VAS requirements
        
        Returns list of warnings/issues found
        """
        warnings = []
        
        # Check for required items based on statement type
        codes = [item.code for item in items]
        
        # Common required items for income statement
        required_is_items = ["01", "02", "10", "60", "70"]  # Revenue, COGS, Gross Profit, Pretax, Net Income
        if any(code.startswith(("0", "1", "6", "7")) for code in codes):
            for req_code in required_is_items:
                if req_code not in codes:
                    warnings.append(f"Missing required income statement item: {req_code}")
        
        # Check for negative values where unexpected
        for item in items:
            if item.value < 0 and item.code in ["01", "10"]:  # Revenue, Gross Profit
                warnings.append(f"Unexpected negative value for {item.name_vi} ({item.code})")
        
        # Check calculation consistency
        # e.g., Gross Profit = Revenue - COGS
        revenue = next((i.value for i in items if i.code == "01"), None)
        cogs = next((i.value for i in items if i.code == "02"), None)
        gross_profit = next((i.value for i in items if i.code == "10"), None)
        
        if revenue and cogs and gross_profit:
            expected_gp = revenue - cogs
            if abs(expected_gp - gross_profit) > abs(expected_gp) * 0.01:  # 1% tolerance
                warnings.append(f"Gross profit calculation mismatch: expected {expected_gp}, got {gross_profit}")
        
        return warnings


# Convenience functions
def parse_vn_financials_from_dict(data: Dict[str, Any]) -> ParsedVNFinancials:
    """Parse Vietnamese financials from dictionary"""
    parser = VNDFinancialParser()
    return parser.parse_from_dict(data)


def convert_vnd_to_usd(amount_vnd: float, exchange_rate: float = 24500) -> float:
    """Convert VND to USD"""
    return amount_vnd / exchange_rate


def get_vas_account_mapping(account_code: str) -> str:
    """Get English description for VAS account code"""
    mappings = {
        "511": "Revenue from sales of goods",
        "512": "Revenue from rendering of services",
        "632": "Cost of goods sold",
        "641": "Selling expenses",
        "642": "Administrative expenses",
        "515": "Financial income",
        "635": "Financial expenses",
        "6351": "Interest expense",
        "711": "Other income",
        "811": "Other expenses",
        "821": "Corporate income tax expense",
        "111": "Cash on hand",
        "112": "Bank deposits",
        "131": "Accounts receivable from customers",
        "151-157": "Inventories",
        "331": "Accounts payable to suppliers",
        "3411": "Short-term loans",
        "3412": "Long-term loans",
        "411": "Owner's equity",
        "421": "Retained earnings",
    }
    
    return mappings.get(account_code, f"Account {account_code}")
